#coding:utf-8

import urllib.request
import lxml.html
from pymarc import Record, Field
from pymarc import MARCReader
import re
import xlwt
import sys,io
import openpyxl
from bs4 import BeautifulSoup
import gzip
import docx
from docx import Document
from io import BytesIO
import pymysql
import pinyin
import datetime
import requests

#改变标准输出的默认编码
#sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='gb18030')

headers = {
    'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    #'Accept-Encoding':'gzip, deflate, sdch',
    'Accept-Encoding':'gzip, deflate',
    'Accept-Language':'zh-CN,zh;q=0.9',
    #'Connection':'keep-alive',
    #'Cookie':'_gscu_413729954=00942062efyg0418; Hm_lvt_2cb70313e397e478740d394884fb0b8a=1500942062',
    #'Host':'opac.nlc.cn',
    'Cookie':'PHPSESSID=0f94e40864d4e71b5dfeb2a8cf392922; Hm_lvt_668f5751b331d2a1eec31f2dc0253443=1542012452,1542068702,1542164499,1542244740; Hm_lpvt_668f5751b331d2a1eec31f2dc0253443=1542246351',
    'Upgrade-Insecure-Requests':'1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3141.7 Safari/537.36 Core/1.53.3226.400 QQBrowser/9.6.11682.400'}


def getHtml(url,num_retries = 5):
    print('Crawling url:',url)
    try:
        request = urllib.request.Request(url,headers=headers)
        response = urllib.request.urlopen(request,timeout=30)
        info = response.info();
        #print(info)
        page_html = ''
        #print(page_html)
        page_html = response.read()

        if info.get('Content-Encoding') == 'gzip':
            buff = BytesIO(page_html)  # 把content转为文件对象
            f = gzip.GzipFile(fileobj=buff)
            page_html = f.read().decode('utf-8')
        else:
            #the_page = response.read().decode('utf-8').encode('utf-8')
            page_html = page_html.decode('utf-8','ignore')
            print(page_html)
        #print(page_html)
        #the_page = response.read().decode('utf-8').encode('utf-8')
        #page_html  = the_page.decode('utf8')
    except Exception as e:
        print('Downloading error:',str(e))
        print('重试次数：', num_retries)
        page_html = None;
        if (num_retries > 0):
            if(hasattr(e, 'code') and 500 <= e.code < 600) :
                return getHtml(url,num_retries - 1)
            else:
                return getHtml(url, num_retries - 1)
        else :
            print('重试次数完毕：',num_retries)
            return page_html
    return page_html


def get_page_html():
    opac_url = 'http://opac.nlc.cn/F'
    opac_url_html = getHtml(opac_url)
    page_tree = lxml.html.fromstring(opac_url_html)
    action = page_tree.xpath('//form[@method="get"]')[0].attrib['action']
    print('Get URL : ' + action)
    return action

def search_isbn(out,url,isbn,row) :
    url_params = url + '?func=find-b&find_code=ISB&request=%s&local_base=NLC01&filter_code_1=WLN&filter_request_1=&filter_code_2=WYR&filter_request_2=&filter_code_3=WYR&filter_request_3=&filter_code_4=WFM&filter_request_4=&filter_code_5=WSL&filter_request_5='
    page_context = lxml.html.fromstring(getHtml(url_params % isbn))
    inputs = page_context.xpath('//input[@id="set_number"]')
    if(len(inputs) > 0) :
        set_number = inputs[0].attrib['value']
        marc_url_params = '?func=full-set-set-body&set_number=%s&set_entry=000001&format=001';
        marc_url_params = marc_url_params % (set_number)
        marc_url = url + marc_url_params
        out_marc(marc_url,out,row)
    print(inputs)

def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    url = get_page_html()
    out = open('C:/Users/dell/Desktop/sciencep20180621_2.mrc', 'wb')
    i = 0
    #for i in range(0,len(sheets)) :
    sheet = wb['Sheet1']
    for row in sheet.rows:
        #if(row[1].value == '没有'):
            #if(row[5].value != '无'):
        search_isbn(out,url,row[3].value,row)
            #else :
            #    i = i + 1
             #   print('i =',i)
    out.close()

def insertMysql(sql):
    #sql = pymysql.escape_string(sql)
    lastid = 0
    db = pymysql.connect(host='localhost',port= 3306,user = 'root',passwd='123456',db='zhiwu',charset='utf8')
    cursor = db.cursor()
    db.escape(sql)
    try:
        #print(sql)
        cursor.execute(sql)
        lastid = db.insert_id();
        db.commit()
    except Exception as e:
        print(e)
        db.rollback()
    cursor.close()
    db.close()
    return lastid

def get_pinyin(str):
    #print(pinyin.get_initial(str,delimiter=''))
    # if isinstance(var_str, str):
    #     if var_str == 'None':
    #         return ""
    #     else:
    #         return pinyin.get(var_str, format='strip', delimiter="")
    # else:
    #     return ''
    if(str is None) :
        return ''
    str = str.strip()
    if(str == 'None' or str == ''):
        return ''
    return pinyin.get(str,format='strip',delimiter=' ')
def get_pinyin_prefix(str):
    #return pinyin.get_initial(str,delimiter='').upper()
    # if isinstance(var_str, str):
    #     if var_str == 'None':
    #         return ""
    #     else:
    #         return pinyin.get_initial(var_str, format='strip', delimiter="").upper()
    # else:
    #     return ''
    if (str is None):
        return ''
    str = str.strip()
    if (str == 'None' or str == ''):
        return ''
    return pinyin.get_initial(str,delimiter='').upper()

def read07_excel(path):
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    sheet = sheet = wb.get_sheet_by_name(sheets[0])
    pid = 0
    shu_id = 0
    is_existed_kes = []
    for row in sheet.rows:
        #print(row[2].value)
        if row[2].value is not None: #种
            print(row[0].value, row[1].value, row[2].value, row[3].value)

            latin = row[1].value
            mings = latin.split(' ')
            latin1 = ''
            zuozhe = []
            if len(mings) > 1:
                print('latin = ',mings[0],mings[1])
                latin1 = mings[0] + ' ' + mings[1]
                shu = mings[0]
                zuozhe = mings[2:len(mings)]
                if shu is not None:
                    if shu not in is_existed_kes:
                        shu_data = get_name_existed(shu)
                        if shu_data is not None:
                            sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`type1`,`create_date`,`del_flag`,`mark`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                            sql = sql % (18, '属', pid, shu, shu_data[5], shu_data[6], shu, get_pinyin(shu_data[6]),
                                 get_pinyin_prefix(shu_data[6]), datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                 1, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0,196)
                            shu_id = insertMysql(sql)
                            is_existed_kes.append(shu)
                sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`type1`,`create_date`,`del_flag`,`mark`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                sql = sql % (26, 'sp.', shu_id, latin1, ' '.join(zuozhe), row[0].value, latin1, get_pinyin(row[0].value),
                         get_pinyin_prefix(row[0].value), datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 1,
                         datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0,196)
                cid = insertMysql(sql)

                if (row[2].value is not None):
                    sql1 = "insert into tb_spdesc (`spcid`,`splatin2`,`spdescid`,`spdesc`,`sporderid`,`spaddtime`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s')"
                    sql1 = sql1 % (cid, latin1, 7, row[2].value, 7, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0)
                    insertMysql(sql1)
                if (row[3].value is not None):
                    sql1 = "insert into tb_spdesc (`spcid`,`splatin2`,`spdescid`,`spdesc`,`sporderid`,`spaddtime`,`create_date`,`del_flag`,`isen`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                    sql1 = sql1 % (cid, latin1, 7, row[3].value, 7, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0,1)
                    insertMysql(sql1)
        else:
            print('科：',row[0].value, row[1].value)
            sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`type1`,`create_date`,`del_flag`,`mark`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
            sql = sql % (13, '科', 0, row[1].value, '', row[0].value, row[1].value, get_pinyin(row[0].value),get_pinyin_prefix(row[0].value), datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 1,datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0,196)
            pid = insertMysql(sql)
    wb.close()

def read_xlsx(path):
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    print(sheets)
    sheet = wb.get_sheet_by_name(sheets[0])
    xlsx_rows = sheet.rows;
    wb.close();
    return xlsx_rows

def read07_excel_by(path):
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    sheet = sheet = wb.get_sheet_by_name(sheets[0])
    pid = 0
    shu_id = 0
    ke_id = 0
    is_existed_kes = []
    is_existed_shus = []
    cid = 0
    for row in sheet.rows:
        #print(row[2].value)
        if row[2].value is not None: #种
            print(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value,row[6].value)
            latin = row[2].value
            mings = latin.split(' ')
            if len(mings) > 1:
                print('latin = ',mings[0],mings[1])
                shu = mings[0]
                if row[3].value is not None :
                    if row[3].value not in is_existed_kes:
                        ke_data = get_name_existed(row[3].value)
                        print('科信息：',ke_data)
                        if ke_data is not None:
                            sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`type1`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                            sql = sql % (13, '科', pid, ke_data[4], ke_data[5], ke_data[6], ke_data[4], get_pinyin(row[3].value),
                                         get_pinyin_prefix(row[3].value),
                                         datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                         1, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 196)
                            ke_id = insertMysql(sql)
                            is_existed_kes.append(row[3].value)
                            print(sql)
                if shu is not None:
                    if shu not in is_existed_shus:
                        shu_data = get_latin2_existed(shu)
                        if shu_data is not None:
                            sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`type1`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                            sql = sql % (18, '属', ke_id, shu, shu_data[5], shu_data[6], shu, get_pinyin(shu_data[6]),
                                 get_pinyin_prefix(shu_data[6]), datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                 1, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 196)
                            shu_id = insertMysql(sql)
                            is_existed_shus.append(shu)
                            print(sql)


                sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`type1`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                sql = sql % (26, 'sp.', shu_id, latin, '', row[4].value, latin, get_pinyin(row[4].value),
                         get_pinyin_prefix(row[4].value), datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 1,
                         datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 196)
                cid = insertMysql(sql)
                print(sql)
                if (row[5].value is not None):
                    sql1 = "insert into tb_spdesc (`spcid`,`splatin2`,`spdescid`,`spdesc`,`sporderid`,`spaddtime`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s')"
                    sql1 = sql1 % (cid, latin, 7, row[5].value, 7, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0)
                    print(sql1)
                    insertMysql(sql1)
                if (row[6].value is not None):
                    sql1 = "insert into tb_spdesc (`spcid`,`splatin2`,`spdescid`,`spdesc`,`sporderid`,`spaddtime`,`create_date`,`del_flag`,`isen`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                    sql1 = sql1 % (cid, latin, 4, row[6].value, 4, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0,1)
                    print(sql1)
                    insertMysql(sql1)
        #if row[0].value == '30':break
        # else:
        #     print('科：',row[0].value, row[1].value)
        #     sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`type1`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
        #     sql = sql % (13, '科', 0, row[1].value, '', row[0].value, row[1].value, get_pinyin(row[0].value),get_pinyin_prefix(row[0].value), datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 1,datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 3)
        #     pid = insertMysql(sql)
    wb.close()

def get_latin2_existed(name):
    db = pymysql.connect(host='localhost',port= 3306,user = 'root',passwd='123456',db='zhiwu',charset='utf8')
    cursor = db.cursor()
    sql = 'select * from tb_classsys where classsys_latin2=\'%s\''
    sql = sql % name
    cursor.execute(sql)
    data = cursor.fetchone()
    #print(data)
    cursor.close()
    return data

def update(sql,db_name='zhiwu'):
    db = pymysql.connect(host='localhost', port=3306, user='root', passwd='123456', db=db_name, charset='utf8')
    cursor = db.cursor()
    cursor.execute(sql)
    int_num = cursor.rowcount
    db.commit()
    cursor.close()
    return int_num

def select(sql,db_name='zhiwu',select_one = True):
    db = pymysql.connect(host='localhost', port=3306, user='root', passwd='123456', db=db_name, charset='utf8')
    cursor = db.cursor()
    cursor.execute(sql)
    if select_one:
        data = cursor.fetchone()
    else:
        data = cursor.fetchall()
    cursor.close()
    return data
def select_all(sql,db_name='zhiwu'):
    db = pymysql.connect(host='localhost', port=3306, user='root', passwd='123456', db=db_name, charset='utf8')
    cursor = db.cursor()
    cursor.execute(sql)
    data = cursor.fetchall()
    cursor.close()
    return data

def get_keshu():
    db = pymysql.connect(host='localhost',port= 3306,user = 'root',passwd='123456',db='ke_shu_dict',charset='utf8')
    cursor = db.cursor()
    sql = 'select * from tb_classsys where del_flag=0'
    #sql = sql % name
    cursor.execute(sql)
    data = cursor.fetchall()
    #print(data)
    cursor.close()
    return data
def get_keshu_spdesc(cid):
    db = pymysql.connect(host='localhost',port= 3306,user = 'root',passwd='123456',db='ke_shu_dict',charset='utf8')
    cursor = db.cursor()
    sql = 'select * from tb_spdesc where spcid=%s'
    sql = sql % cid
    cursor.execute(sql)
    data = cursor.fetchall()
    #print(data)
    cursor.close()
    return data

def get_name_existed(name):
    db = pymysql.connect(host='localhost',port= 3306,user = 'root',passwd='123456',db='zhiwu',charset='utf8')
    cursor = db.cursor()
    sql = 'select * from tb_classsys where classsys_cname=\'%s\''
    sql = sql % name
    cursor.execute(sql)
    data = cursor.fetchone()
    #print(data)
    cursor.close()
    return data
def get_foc():
    db = pymysql.connect(host='localhost', port=3306, user='root', passwd='123456', db='zhiwu', charset='utf8')
    cursor = db.cursor()
    sql = 'select * from zhiwu2'
    #sql = sql % name
    cursor.execute(sql)
    data = cursor.fetchall()
    print(data)
    cursor.close()
    return data


def out_marc(marcUrl,out,row) :
    page_detail = getHtml(marcUrl)
    page_context = BeautifulSoup(page_detail,"html.parser")
    trs = page_context.find_all("tr")
    record = Record()
    record.force_utf8 = "true"
    for tr in trs:
        tds = tr.find_all("td")
        fn = tds[0].string.replace("\s$"," ").replace("\xa0"," ")
        fv = tds[1].string.replace("\s$"," ").replace("\xa0"," ").replace("\x1e"," ")
        if(fn.upper() == "LDR" ) :
            record.leader = fv
            continue
        if(fn.upper() == "FMT" ) :
            continue
        if(fn.upper() == "SYS" ) :
            continue
        if(fn.upper() == "CAT" ) :
            continue
        if(fn.upper() == "OWN" ) :
            continue
#         if(fn.upper() == "049" ) :
#             continue;
#         if(fn.upper() == "090" ) :
#             continue;
#         if(fn.upper() == "096" ) :
#             continue;
        pattern = re.compile('\\|([A-Za-z0-9_]?)\\s')
        m1 = pattern.findall(fv)
        ind = []
        jj = 0
        for l in m1:
            ind.append(jj)
            jj += 1
        m = re.split(r'\|([A-Za-z0-9_]?)\s',fv)
        _m = []
        for s in m:
            if(s) :
                _s = s.replace("\s$"," ").strip()
                if(_s) :
                    _m.append(_s)
        #print("--------------------------------")
        if(fn.isdigit() or len(_m) > 1) :
            if(fn.isdigit() and len(_m) == 1):
                field = Field(tag=fn,data = _m[0])
            else :
                if(len(fn) == 3) :
                    field = Field(tag=fn,indicators=[' ',' '],subfields=_m)
                elif(len(fn) == 4):
                    field = Field(tag=fn[0:3],indicators=[fn[3:4],' '],subfields=_m)
                else:
                    field = Field(tag=fn[0:3],indicators=[fn[3:4],fn[4:5]],subfields=_m)
        else :
            if(len(_m) == 1) :
                field = Field(tag=fn,indicators=[' ',' '],subfields=_m)
            else :
                field = Field(tag=fn,data = _m[0])

        #if(field.tag == '010'):
        #    isbn = field.subfields[1]
        record.add_field(field)
    #field = Field(tag='856',indicators=['4',' '],subfields=['u',"http://book.sciencereading.cn/shop/book/Booksimple/show.do?id="+str(row[0].value)])
    #record.add_field(field)
    #field = Field(tag='999',indicators=[' ',' '],subfields=['a',str(row[0].value)])
    #record.add_field(field)
    print('当前记录 = ' +'---' + record.as_json())
    out.write(record.as_marc())

def get_text_docx():
    file = docx.Document("C:\\Users\\dell\\Desktop\\高等植物三卷.docx")
    i = 1
    j = 0
    wb = xlwt.Workbook()
    ws = wb.add_sheet('中国高等植物彩色图鉴正文内容-第三卷', cell_overwrite_ok=True)
    ws.write(0, 0, '物种中文名')
    ws.write(0, 1, '物种拉丁名')  # 科-中文名
    ws.write(0, 2, '正文内容')
    ws.write(0, 3, '正文英文内容')
    ke = False
    for p in file.paragraphs:
        #if i > 20 :break
        #print('--------------------')d
        #if p.text.strip() == '':break

        if p.text.strip() == '' :
            continue
        if j%4 == 0:
            j = 0
            i = i + 1


        print('----------',i, j)
        ws.write(i, j, p.text.strip())

        if ke is True:
            j = 0
            #i = i + 1
            ke = False
        else :
            j = j + 1
        print(p.text,'---',p.style.name)
        #print(run.bold for run in p.runs)
        #if p.style.name == '种-英文' :
        for run in p.runs:
            if run.bold :
                print(run.text,run.bold)
            #print(run.bold)
        #print('--------------------')
        #j = j + 1
        if p.text.strip().endswith('科'):
            ke = True

    wb.save("C:/Users/dell/Desktop/高等三卷.xls")



def get_content(url='http://www.efloras.org/',cralw_url='http://www.efloras.org/browse.aspx?flora_id=2&page=%s',pages=2):
    for i in range(1,pages+1):
        cralw_url_i = cralw_url % (str(i))
        info = getHtml(cralw_url_i)
        #print(info)
        page_context = BeautifulSoup(info, "html.parser")
        divs = page_context.find_all(id='ucFloraTaxonList_panelTaxonList')
        #print(divs)
        if len(divs) > 0:
            div = divs[0]
            table = div.find_all('table')[0]
            #print(table)
            trs = table.find_all('tr')
            #print(trs)
            for tr in trs:
                tds = tr.find_all('td')
                if len(tds) == 5:
                    print(tds[0].text,tds[1].text,tds[2].text,tds[3].text,tds[4].text)
                #if tds[1].fina_all('a') is not None:
                    ke_urls = tds[1].select('a[href]')
                    print(ke_urls)
                    if len(ke_urls) > 0:
                        ke_url = ke_urls[0].get('href');
                        print('ke_url :',ke_url)
                        ke_context = getHtml(url)
                        #print(ke_context)
                        ke_context_soup = BeautifulSoup(ke_context, "html.parser")
                        table_ke = ke_context_soup.find_all('table',id='footerTable')
                        print(table_ke)
                    shu_urls = tds[3].select('a[href]')
                    print(shu_urls)
                    if len(shu_urls) > 0:
                        print('shu_url :',shu_urls[0].get('href'))
                # else:
                #     pages = tds[1].find_all('a',title='Page 2')
                #     print(pages)
                #     if len(pages) > 0:
                #      for page in pages:
                #          page_c = BeautifulSoup(getHtml('http://www.efloras.org/' + page.get('href')), "html.parser")
                #         page_context_c = BeautifulSoup(page_c, "html.parser")
            # for td in tds:
            #     text = td.text;
            #     print(text)
    # lis = uls[0].find_all('tr')
    # for li in lis:
    #     #print(li.select('td'))
    #     links = li.select('td')
    #     #print(links.text)
    #     for td in links:
    #         print(td.text)
    #     #content = getHtml('http://www.jetsen.com.cn' + links)
    #     #print(content)

def get_ke_context(url):
    volume_content = {};
    ke_context = getHtml(url)
    volume_content['url'] = url
    volume_content['taxon_id'] = get_max_number(url)

    ke_context_soup = BeautifulSoup(ke_context, "html.parser")
    table_ke = ke_context_soup.find_all('table', id='footerTable')
    tds = table_ke[0].select('td[style]')
    #print(tds[0].text)# 科所在的卷册、页码等
    volume_content['volume_title'] = tds[0].text
    div_context = ke_context_soup.find_all('div', id='panelTaxonTreatment')
    #print(div_context[0].find_all(re.compile("^image")))
    #print('正文内容：',div_context[0].prettify())
    foc_taxon_chain = ke_context_soup.select_one('span[id="lblTaxonChain"]')
    #print(foc_taxon_chain)
    parent_links = foc_taxon_chain.find_all('a')
    if parent_links:
        parent_link = parent_links[len(parent_links)-1].get('href')
        volume_content['parent_taxon_id'] = get_max_number(parent_link)
    volume_list = foc_taxon_chain.find_all('a', href=re.compile("volume_id"), recursive=False)
    if len(volume_list) == 1:
        volume_content['volume_id'] = get_max_number(volume_list[0].get('href'))
        volume_content['volume'] = volume_list[0].text
    span = div_context[0].find_all('span',id='lblTaxonDesc')[0]

    #print('正文内容：', span.prettify())
    #print(span.prettify())

    #####################获取有image图片信息的部分内容################
    image_table = span.select_one('table')
    if image_table:
        image_table_tr_list = image_table.find_all('tr')
        for image_table_tr in image_table_tr_list:
            image_table_td_list = image_table_tr.find_all('td')
            for image_table_td in image_table_td_list:
                if image_table_td.a:
                    #print('图片连接：',image_table_td.select_one('a').img.get('src'))          ##获取图片的链接\
                    image_link = image_table_td.a.img.get('src')
                    #print('图片连接：', image_link)

                    #download_file(image_link,'F:\FloraData\images\\' + str(get_max_number(image_link)) + '.jpg')
                if image_table_td.a.next_sibling :
                    print('当前物种的拉丁名及链接等：',image_table_td.a.next_sibling.get('href'),image_table_td.a.next_sibling.text)
                if image_table_td.a.next_sibling.next_sibling:
                    print('Credit:',image_table_td.a.next_sibling.next_sibling.small.text)
        image_table.extract()
    ###############################################################
    #print(span.b.next_siblings)
    latin_name_object = []
    for wuzh in span.next_element.next_siblings:
        if wuzh.name == 'p':continue
        if wuzh.name == 'a': #表示直接跳转下个物种，类似 See Isoëtaceae # http://www.efloras.org/florataxon.aspx?flora_id=2&taxon_id=20790
            latin_name_object = []
            latin_name_object.append(wuzh)
            break
        #if wuzh is not None : print('---------:', wuzh.name,repr(wuzh))
        if wuzh.string.strip('\n\r '):
            latin_name_object.append(wuzh)
        #else:
        #    print(repr(wuzh).strip(['\n', ' ', '\r\n']))
    print(latin_name_object)
    if len(latin_name_object) > 1:
        if latin_name_object[0].name is None: #如果第一个字符串是类似1.，7a,... 则表示序号
            volume_content['xuhao'] = latin_name_object[0].string.strip('\n\r ')
        else:
            volume_content['xuhao'] = ''
        if latin_name_object[len(latin_name_object)-1].name is None : #如果最后一个字符串是类似(Blume) Tagawa, Acta Phytotax. Geobot. 7: 83. 1938.则表示文献
            volume_content['latin_name'] = ' '.join(list(latin.string.strip('\n\r ') for latin in latin_name_object[1:len(latin_name_object)-1] ))
        else:
            volume_content['latin_name'] = ' '.join(list(latin.string.strip('\n\r ') for latin in latin_name_object[1:]))
    else:
        volume_content['xuhao'] = ''
        volume_content['latin_name'] = ' '.join(list(latin.string.strip('\n\r ') for latin in latin_name_object))
    #volume_content['xuhao'] = latin_name[0]
    #print(span.b.next_sibling) #当前物种信息的物种拉丁名
    #print(span.b.find_next_sibling("p").contents[0].strip())
    volume_content['latin_name_full'] = span.b.next_sibling.strip()
    #print(span.b.find_next_sibling("p"))
    #print('-----------------------')
    #print(span.b.find_next_sibling("p").contents[0])
    zh_name_and_pinyin = span.b.find_next_sibling("p").contents[0]
    if is_all_zh(zh_name_and_pinyin):   #含有中文
        print('#######################')
        print(zh_name_and_pinyin.split(' ')[0].strip())
        print(' '.join(zh_name_and_pinyin.split(' ')[1:]))
    #print(re.sub('[A-Za-z0-9\!\%\[\]\,\。\(\)]', '', zh_name_and_pinyin))
    #print(' '.join(re.findall(r'[A-Za-z\(\)]+', zh_name_and_pinyin)))
        volume_content['zh_name'] = zh_name_and_pinyin.split(' ')[0].strip()
        volume_content['zh_name_pinyin'] = ' '.join(zh_name_and_pinyin.split(' ')[1:]).strip()
    else:
        volume_content['zh_name'] = ''
        volume_content['zh_name_pinyin'] = zh_name_and_pinyin.strip()
    #authors = span.b.find_next_sibling("p").p.next_element #获取下面一个直接字符串
    spdesc_p_list = span.b.find_next_sibling("p").p
    #print('##############################################')
    #print(spdesc_p_list)
    #print('##############################################')
    #print(spdesc_p_list.find_all('a',recursive=False))
    authors_list = []
    authors_id_list = []
    for author in spdesc_p_list.find_all('a',recursive=False):
        #print(author.text,author.get('href'))
        authors_list.append(author.text)
        authors_id_list.append(str(get_max_number(author.get('href'))))
    volume_content['authors'] = ';'.join(authors_list)
    volume_content['authors_id'] = ';'.join(authors_id_list)

    #print(authors.find_all('p',recursive=False)[0].prettify())
    spdescs = spdesc_p_list.find_all('p',recursive=False)
    #print(spdescs)
    print('##############################################')
    if len(spdescs) > 0:
        specs_context = ''
        table = spdescs[0].select_one('table')
        if table is not None:
            #print(table.find_all('a'))

            for s in table.next_sibling.next_sibling.strings:
                #print(repr(s),type(s),s.parent.name=='i')
                if s.parent.name == 'i':
                    specs_context = specs_context + '<i>' + s.strip('\n') + '</i>'
                else:
                    specs_context = specs_context + s.strip('\n')
        else :
            #print(spdescs[0].strings)
            for s in spdescs[0].strings:
                #print(s)
                if s.parent.name == 'i':
                    specs_context = specs_context + '<i>' + s.strip('\n') + '</i>'
                else:
                    if s.parent.name == 'b':
                        specs_context = specs_context + '<b>' + s.strip('\n') + '</b>'
                    specs_context = specs_context + s.strip('\n')
            #print(specs_context.strip())
        #print('##############################################')
        #print(specs_context.strip())#获取正文内容
        volume_content['content'] = specs_context.strip()
    #volume_content['create_date'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # sql = "insert into volume_content (`content`,`create_date`,`del_flag`) values ('%s','%s','%s')"
        # sql = sql % (specs_context.strip(), datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0)
        # pid = insertMysql(sql)
    #print('##############################################')
    print(volume_content)
    wuzhong_detail_sql(volume_content)
    table_jiansuobiao = div_context[0].find_all('table',id='tableKey') #获取检索表的内容
    if len(table_jiansuobiao) > 0:
        trs_jiansuobiao = table_jiansuobiao[0].find_all('tr')
        table_jsb = trs_jiansuobiao[1].find_all('table')
        if len(table_jsb) > 0:
            trs_jsb = table_jsb[0].find_all('tr')
            for tr in trs_jsb:
                tds_jsb = tr.find_all('td')
                tds_jxb_cs = tds_jsb[3].contents;
                goto_no = ''
                goto_id = ''
                for tds_jxb_c in tds_jxb_cs:
                    #print(tds_jxb_c.name)
                    if tds_jxb_c.name == 'a':
                        tds_jxb_c_href = tds_jxb_c.get('href')
                        tds_jxb_c_s = tds_jxb_c.string;
                        if tds_jxb_c_s is not None:
                            #print(tds_jxb_c)
                            goto_id = tds_jxb_c_href + '='+ tds_jxb_c_s
                        else:
                            goto_id = tds_jxb_c_href
                    else :
                        goto_no = tds_jxb_c
                #print(tds_jsb[0].text,tds_jsb[1].text,tds_jsb[2].text,goto_no,goto_id)
    ############################################################################################
    ###lower_taxa_ul = div_context[0].select_one('ul')#获取当前物种的下级物种信息
    ###print(lower_taxa_ul)
    # if lower_taxa_ul is not None:
    #     for li in lower_taxa_ul.find_all('li'):
    #         lower_taxa_a = li.select_one('a')
    #         #print(lower_taxa_a.get('href'),lower_taxa_a.b.string,lower_taxa_a.b.next_sibling)
    ############################################################################################
    related_objects = div_context[0].select_one('span[id="lblObjectList"]')
    #print(related_objects)
    if related_objects is not None:
        related_objects_trs = related_objects.find_all('tr')
        #print(related_objects_trs)
        for related_objects_tr in related_objects_trs:
            related_objects_tds = related_objects_tr.find_all('td')
            if len(related_objects_tds) == 2:
                related_objects_td_li = related_objects_tds[0].li
                if related_objects_td_li is not None:
                    li_a = related_objects_td_li.a
                    print(li_a.text,li_a.get('href'))
                else:
                    print(related_objects_tds[0].text)
                print(related_objects_tds[1].text)
            else:
                print('采集错误')

def get_foc_vol_list(url='http://www.efloras.org/index.aspx'):
    context = getHtml(url)
    context_soup = BeautifulSoup(context, "html.parser")
    span = context_soup.find_all('span',id='lblFloraList')
    url_list = []
    #print(span)
    if len(span) > 0:
        ul_list = span[0].find_all('ul')
        li_list = ul_list[2].find_all('li') #FOC在ul_list的第三个位置
        a_list = li_list[1].find_all('a')
        print(a_list)
        for a in a_list[1:]:  #a_list[1:]:
            a_href = a.get('href')
            print(' Volume :',a.text)
            #sql = "insert into volume (`url`,`volume_id`,`volume_no`,`create_date`,`create_by`,`del_flag`) values ('%s','%s','%s','%s','%s','%s')"
            if a_href is not None:
                url_list.append('http://www.efloras.org/' + a_href)
                volume_id = re.sub("\D", "", a_href.split('&')[0])
                print('volume_id',volume_id)
                #sql = sql % ('http://www.efloras.org/' + a_href, volume_id, a.text,
                #            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'luoxuan', '0')
                #insertMysql(sql)
            #else:
                #sql = sql % ('','',a.text,datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),'luoxuan','0')

    else:
        print('未找到')

    return url_list

def get_foc_volume_list(volumes):
    url_list = []
    for vol in volumes:
        context = getHtml(vol)
        if context is None:continue
        context_soup = BeautifulSoup(context, "html.parser")
        div = context_soup.find_all('div', id='ucFloraTaxonList_panelTaxonList')
        volumeInfo = context_soup.select_one('span[id="ucVolumeInfo_lblVolumeInfo"]')
        volume_map = []
        if volumeInfo is not None:
            volumeInfo_table_trs = volumeInfo.table.find_all('tr')
            if len(volumeInfo_table_trs) > 0:
                for volumeInfo_table_tr in volumeInfo_table_trs:
                    volumeInfo_table_tds = volumeInfo_table_tr.find_all('td')
                    if len(volumeInfo_table_tds) == 2:
                        volume_map.append(volumeInfo_table_tds[1].text)
                    else:
                        volume_map.append('')
        if len(volume_map) != 5:
            for i in range(5-len(volume_map)): volume_map.append('')
        print(volume_map)
        if len(div) > 0:
            tr_list = div[0].find_all('tr',class_='underline')
            #print(tr_list)
            for tr in tr_list[2:]:
                #print(tr)
                td_list = tr.find_all('td')
                wuzhong_detail_link_a = td_list[1].select_one('a')
                wuzhong_detail_link = ''
                if wuzhong_detail_link_a:wuzhong_detail_link = wuzhong_detail_link_a.get('href')
                lower_taxa_link_a = td_list[3].select_one('a')
                lower_taxa_link = ''
                if lower_taxa_link_a:lower_taxa_link = lower_taxa_link_a.get('href')
                print('Taxon_Id :',td_list[0].text,'Latin Name :',td_list[1].text,'Url :',wuzhong_detail_link,
                      'Zh_Name:',td_list[2].text,'Lower_Taxa :',td_list[3].text,'Lower_Taxa_link :',lower_taxa_link)
                ke_url = td_list[1].select_one('a').get('href')
                get_ke_context('http://www.efloras.org/' + ke_url)
                sql = "insert into volume_ke (`parent_taxon_id`,`type`,`type_name`,`taxon_id`,`taxon_name`,`accepted_name`,`accepted_name_url`,`accepted_name_cn`,`lower_taxa`,`lower_taxa_url`,`volume_no`,`volume_name`,`title`,`families`,`genera`,`speces`,`online_date`,`create_date`,`create_by`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                sql = sql % (get_max_number(vol),'','',td_list[0].text.strip(), '',td_list[1].text.strip(), wuzhong_detail_link,td_list[2].text.strip(), td_list[3].text.strip(),lower_taxa_link, get_max_number(vol),'', volume_map[0], volume_map[1],volume_map[2],volume_map[3],volume_map[4],datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),'luoxuan',0)
                if lower_taxa_link :
                    url_list.append('http://www.efloras.org/' + lower_taxa_link)
                insertMysql(sql)
                #
                # sql = sql % (td_list[0].text, td_list[1].text, td_list[1].select_one('a').get('href'), td_list[2].text, td_list[3].text,
                #              td_list[3].select_one('a').get('href'), re.sub("\D", "", vol.split('&')[0], '', '','','','','','',
                #                 datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),'luoxuan')
                # insertMysql(sql)
            get_foc_volume_list(url_list)

        else:
            print('无法找到')
        volume_related_links_table = context_soup.find_all('table', id='ucVolumeResourceList_dataListResource')
        #print(volume_related_links_table)
        if len(volume_related_links_table) > 0:
            #print(volume_related_links_table[0])
            volumes_relateds = volume_related_links_table[0].find_all('tr',recursive=False) #搜索当前节点的直接子节点
            if len(volumes_relateds) > 0:
                #print(volumes_relateds)
                for volume in volumes_relateds[1:]:
                    #print(volume,'------------')
                    trs=volume.find_all('tr')
                    if len(trs) > 0:
                        #print(trs[0])
                        tds = trs[0].find_all('td')
                        if len(tds) > 1:
                            #print(tds)
                            #print('--------------------------')
                            a = tds[0].select_one('a')
                            href = a.get('href')
                            print('--------',a.text,' ',href)
                            print('=====',tds[1].text)
                            sql1 = "insert into volume_related_links (`taxid`,`type`,`url`,`title`,`resource_type`,`files`,`create_date`,`create_by`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                            if tds[1].text.strip() == 'PDF':
                                paths = href.split('/')
                                print(paths)
                                #download_file(href,'f://FloraData//' + paths[len(paths)-1])
                                #sql1 = sql1 % (vol.split('&')[0]),tds[1].text,href,a.text,paths[len(paths)-1],datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),'luoxuan',0)
                                sql1 = sql1 % (re.sub("\D", "", vol.split('&')[0]),tds[1].text,href,a.text,'PDF',paths[len(paths)-1],datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),'luoxuan',0)
                            else: #tds[1].text == 'Treatment'
                                #get_ke_context(href)
                                sql1 = sql1 % (re.sub("\D", "", vol.split('&')[0]),tds[1].text,href,a.text,'',tds[1].text,
                                               datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'luoxuan', 0)
                            #print(insertMysql(sql1))
        else:
            print('无法找到volume_related_links')

def get_max_number(str): #获得连接中最大的数字
    return max(list(map(int,re.findall(r"\d+\.?\d*",str))))
def is_all_zh(s): #是否含有中文
    for c in s:
        if not ('\u4e00' <= c <= '\u9fa5'):
            return False
    return True

def wuzhong_detail_sql(volume_content):
    sql = "insert into volume_content (`url`,`content`,`taxon_id`,`parent_taxon_id`,`xuhao`,`latin_name`,`latin_name_full`,`zh_name`,`zh_name_pinyin`,`authors`,`authors_id`,`volume_id`,`volume`,`volume_title`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
    sql = sql % (volume_content['url'],volume_content['content'],volume_content['taxon_id'],volume_content['parent_taxon_id'],volume_content['xuhao'],volume_content['latin_name'],
                 volume_content['latin_name_full'],volume_content['zh_name'],volume_content['zh_name_pinyin'],volume_content['authors'],volume_content['authors_id'],
                 volume_content['volume_id'],volume_content['volume'],volume_content['volume_title'],datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0)
    #13, '科', 0, row[1].value, '', row[0].value, row[1].value, get_pinyin(row[1].value), get_pinyin_prefix(row[1].value),
    #datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 1, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 2)
    pid = insertMysql(sql)
    return pid

def download_file(url,path):
    print('Download file:',url,path)
    r = requests.get(url)
    with open(path, "wb") as code:
        code.write(r.content)

def deal_keshu():
    rows = get_keshu()
    for row in rows:
        print(row)
        spdescs_rows = get_keshu_spdesc(row[0])
        sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`type1`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
        sql = sql % (row[1], row[2], row[3], row[4], row[5], row[6], row[8], row[10], row[11], row[14], row[16],
                     datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 4)
        cid = insertMysql(sql)
        print(sql)
        for spdescs_row in spdescs_rows:
            sql1 = "insert into tb_spdesc (`spcid`,`splatin2`,`spdescid`,`spdesc`,`sporderid`,`spaddtime`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s')"
            sql1 = sql1 % (cid, spdescs_row[2], spdescs_row[3], spdescs_row[4], spdescs_row[5], spdescs_row[6],
                           datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 4)
            print(sql1)
            insertMysql(sql1)

def deal_banquan():
    rows = read_xlsx('C:/Users/dell/Desktop/中国植物志版权页完整版.xlsx')
    sql = "insert into tb_book (`book_name`,`publish_date`,`isbn`,`issue`,`issue_small`,`author`,`price`,`responsible_editor`,`series_name`,`volume_name`,`fascicle_no`,`men`,`gang`,`relates_ke`,`create_date`,`create_by`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"

    for row in rows:
        if row[1] == 'None': continue
        print(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value,
              row[7].value, row[8].value, row[9].value)
        sql_insert = sql % (
        row[3].value, row[0].value, row[1].value, row[6].value, row[2].value, row[4].value, row[5].value, row[7].value,
        row[8].value, row[9].value, row[10].value, row[13].value, row[14].value, row[12].value,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'lx', 0)
        if row[0].value is not None or row[0].value != '初版时间':
            print(sql_insert)
            insertMysql(sql_insert)
def get_parents_classsys(book_id,pid):
    datas = select(
        'select * from tb_classsys where classsys_par = %s and create_date is null and del_flag = 0 and create_by is null' % pid,
        'zhiwu', False)
    if datas is not None:
        ids = []
        for data in datas:ids.append(str(data[0]))
        print(ids)
        if len(ids):
            sql = 'update tb_classsys set mark=%s where id in (%s)' % (str(book_id), ','.join(ids))
            print(sql)
            int_num = update(sql)
            print(int_num)
            for id in ids:
                get_parents_classsys(book_id,id)
        else:
            return
    else:
        return
def zhiwu_book_relates():
    # book_datas = select('select * from tb_book where relates_ke is not null','zhiwu',False)
    book_datas = select('select * from tb_book where id=147', 'zhiwu', False)
    # print(book_datas)
    for book in book_datas:
        book_id = book[0]
        relates_ke = book[14]
        print(book_id, relates_ke)
        if relates_ke is not None:
            ke_arr = relates_ke.split('、')
        else:
            ke_arr = []
            # ke_arr.append('裸子植物门')
        print(ke_arr)
        for ke in ke_arr:
            data = select(
                'select * from tb_classsys where classsys_cname = \'%s\' and create_date is null and del_flag = 0 and create_by is null' % ke,
                'zhiwu')
            if data is not None:
                print(data)
                data_id = data[0]
                int_num = update('update tb_classsys set mark=%s where id=%s' % (book_id, data_id))
                print(int_num)
                get_parents_classsys(book_id, data_id)
def deal_with_gdzwcstj():
    foc_datas = select('select * from cstj where del_flag is null', 'zhiwu', False)
    for data in foc_datas:
        print(data[8], data[9])
        latin2 = data[9]
        zhong = select(
            'select * from tb_classsys where mark in (188,189,190,191,192,193,194) and classsys_latin2 = \'%s\'' % pymysql.escape_string(
                data[9].strip()), 'zhiwu', True)
        if zhong is not None:
            # print(zhong[0])
            int_num = update('update cstj set cid = %s,del_flag = 0 where id = %s' % (zhong[0], data[11]))
            print(zhong, int_num)
def deal_with_foc():
    is_existed_data = [];
    data_count = select('select count(*) from zhiwu2', 'zhiwu', True)
    page_count = data_count[0]
    page_size = 1000
    page_index = page_count // page_size
    if page_count % page_size != 0 :
        page_index = page_index + 1
    for index in range(16,page_index+1):
        sql = 'select id,juan,xueming,zhwenming,ke,shu,zaipei,yisheng,hei,ji,liao,meng,gan,ning,qing,xin,zang,jing,jin,ji1,lu,jin1,shan,yu,e,xiang,dian,gui,chuan,su,wan,zhe,fu,yu1,gan1,gang,ao,gui1,min,yue,tai,qiong from zhiwu2 order by id limit %s,%s'
        sql = sql % (index*page_size,page_size)
        foc_datas_rows = select(sql,'zhiwu', False)
        # print(foc_datas)
        for foc_datas in foc_datas_rows:
            cid = 0
            shuid = 0
            ke_latin = foc_datas[4]
            shu_zh_name = foc_datas[5]
            foc_ke = select('select * from tb_classsys where classsys_latin2 = \'%s\' and del_flag=5 and classsys_classid=13' % pymysql.escape_string(ke_latin), 'zhiwu', True)
            if foc_ke is not None:
                cid = foc_ke[0]
                print('ke_id = ',cid)
            #if ke_latin not in is_existed_data:
            else:
                ke = select('select * from tb_classsys where classsys_latin2 = \'%s\'' % pymysql.escape_string(ke_latin), 'zhiwu', True)
                if ke is not None:
                    ke_sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`mark`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                    ke_sql = ke_sql % (ke[1], ke[2], '0', ke[4], ke[5], ke[6], ke[8], ke[10], ke[11],
                                       datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), foc_datas[1],
                                       datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 5)
                    cid = insertMysql(ke_sql)
                    print(ke_sql)
                    print('ke_id = ',cid)
                else:
                    ke_sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`mark`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                    ke_sql = ke_sql % (13, '科', 0, ke_latin, '', '', ke_latin, '', '',
                                       datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), foc_datas[1],
                                       datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 5)
                    cid = insertMysql(ke_sql)
                    print(ke_sql)
                    print('ke_id = ',cid)
                #is_existed_data.append(ke_latin)
            foc_shu = select('select * from tb_classsys where classsys_cname = \'%s\' and del_flag=5 and classsys_classid=18' % shu_zh_name,'zhiwu', True)
            if foc_shu is not None:
                shuid = foc_shu[0]
                classsys_par = foc_shu[3]
                classsys_par_data = select('select * from tb_classsys where id = %s and del_flag=5' % classsys_par,
                                           'zhiwu', True)
                if classsys_par_data is not None:
                    if classsys_par_data[0] != cid:
                        update('update tb_classsys set classsys_par = %s where id=%s' % (cid, foc_shu[0]))
            #if shu_zh_name not in is_existed_data:
                print('shu_id = ', shuid)
            else:
                shu = select('select * from tb_classsys where classsys_cname = \'%s\'' % shu_zh_name, 'zhiwu', True)
                if shu is not None:
                    shu_sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`mark`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                    shu_sql = shu_sql % (shu[1], shu[2], cid, shu[4], shu[5], shu[6], shu[8], shu[10], shu[11],
                                         datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), foc_datas[1],
                                         datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 5)
                    shuid = insertMysql(shu_sql)
                    print(shu_sql)
                    print('shu_id = ', shuid)
                else:
                    shu_sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`mark`,`create_date`,`del_flag`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                    shu_sql = shu_sql % (
                    18, '属', cid, '', '', shu_zh_name, '', get_pinyin(shu_zh_name), get_pinyin_prefix(shu_zh_name),
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), foc_datas[1],
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 5)
                    shuid = insertMysql(shu_sql)
                    print(shu_sql)
                    print('shu_id = ', shuid)
                    #print(shu)
                #is_existed_data.append(shu_zh_name)
            foc_zhong = select(
                'select * from tb_classsys where classsys_latin2 = \'%s\' and del_flag=5' % pymysql.escape_string(foc_datas[2]),
                'zhiwu', True)
            if foc_zhong is not None:
                classsys_par = foc_zhong[3]
                classsys_par_data = select('select * from tb_classsys where id = %s and del_flag=5' % classsys_par,'zhiwu', True)
                if classsys_par_data is not None:
                    if classsys_par_data[0] != shuid :
                        update('update tb_classsys set classsys_par = %s where id=%s' % (shuid,foc_zhong[0]))
                print(foc_zhong)
            else:
                zhong_sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`mark`,`create_date`,`del_flag`,`zhiwuid`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                zhong_sql = zhong_sql % (26, 'sp.', shuid, foc_datas[2], '', foc_datas[3], foc_datas[2], get_pinyin(foc_datas[3]),get_pinyin_prefix(foc_datas[3]),datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), foc_datas[1],datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 5, foc_datas[0])
                wuzhongid = insertMysql(zhong_sql)
                print('wuzhong_id = ',wuzhongid)
                foc_id = foc_datas[0]
                foc_data = select('select * from tb_classsys where zhiwuid=%s' % foc_id)
                if foc_data is not None:
                    update('update tb_spdesc set spcid=%s,update_by = \'lx\' where spcid = %s and isen = 1' % (
                    wuzhongid, foc_data[0]))
                    # print(foc_data)
                shengfens = [];
                # for d in foc_datas[6:]:
                # print(d)
                if foc_datas[6] is not None:
                    update('update zindexinfo set rid=%s where rid = %s and type=3 ' % (wuzhongid, foc_data[0]))
                    # print(foc_datas[6])
                if foc_datas[7] is not None:
                    update('update zindexinfo set rid=%s where rid = %s and type=4 ' % (wuzhongid, foc_data[0]))
                    update('update zindexinfo set rid=%s where rid = %s and type=2 ' % (wuzhongid, foc_data[0]))
                    # print(foc_datas[7])

                if foc_datas[8] == '1': shengfens.append('黑龙江')
                if foc_datas[9] == '1': shengfens.append('吉林')
                if foc_datas[10] == '1': shengfens.append('辽宁')
                if foc_datas[11] == '1': shengfens.append('内蒙古')
                if foc_datas[12] == '1': shengfens.append('甘肃')
                if foc_datas[13] == '1': shengfens.append('宁夏')
                if foc_datas[14] == '1': shengfens.append('青海')
                if foc_datas[15] == '1': shengfens.append('新疆')
                if foc_datas[16] == '1': shengfens.append('西藏')
                if foc_datas[17] == '1': shengfens.append('北京')

                if foc_datas[18] == '1': shengfens.append('天津')
                if foc_datas[19] == '1': shengfens.append('河北')
                if foc_datas[20] == '1': shengfens.append('山东')
                if foc_datas[21] == '1': shengfens.append('山西')
                if foc_datas[22] == '1': shengfens.append('陕西')
                if foc_datas[23] == '1': shengfens.append('河南')
                if foc_datas[24] == '1': shengfens.append('湖北')
                if foc_datas[25] == '1': shengfens.append('湖南')
                if foc_datas[26] == '1': shengfens.append('云南')
                if foc_datas[27] == '1': shengfens.append('贵州')

                if foc_datas[28] == '1': shengfens.append('四川')
                if foc_datas[29] == '1': shengfens.append('江苏')
                if foc_datas[30] == '1': shengfens.append('安徽')
                if foc_datas[31] == '1': shengfens.append('浙江')
                if foc_datas[32] == '1': shengfens.append('上海')
                if foc_datas[33] == '1': shengfens.append('重庆')
                if foc_datas[34] == '1': shengfens.append('江西')
                if foc_datas[35] == '1': shengfens.append('香港')
                if foc_datas[36] == '1': shengfens.append('澳门')
                if foc_datas[37] == '1': shengfens.append('广西')

                if foc_datas[38] == '1': shengfens.append('福建')
                if foc_datas[39] == '1': shengfens.append('广东')
                if foc_datas[40] == '1': shengfens.append('台湾')
                if foc_datas[41] == '1': shengfens.append('海南')
                difangs = ';'.join(shengfens)
                index_info_sql = "insert into zindexinfo (`type`,`indexinfo`,`rid`,`rtype`,`siteid`,`create_by`,`create_date`,`del_flag`,`update_by`,`update_date`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                index_info_sql = index_info_sql % (
                1, difangs, wuzhongid, 1, 1, 'lx', datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0, 'lx',
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                insertMysql(index_info_sql)
        print('page : ',index*page_size,page_size)
    # for data in foc_datas:

def deal_with_gdcstj(bookid,latin_name,pid):
    data = select('select * from tb_classsys where mark=%s and classsys_latin2=\'%s\'' % (bookid,latin_name),'zhiwu',True)
    if data is None:  ##原有书籍中无法找该物种信息
        shu_data = select('select * from tb_classsys where classsys_latin2=\'%s\'' % (latin_name), 'zhiwu',
                      True)
        sql = "insert into tb_classsys (`classsys_classid`,`classsys_class`,`classsys_par`,`classsys_latin`,`classsys_author`,`classsys_cname`,`classsys_latin2`,`classsys_pinyin`,`classsys_py`,`classsys_addtime`,`type1`,`create_date`,`del_flag`,`mark`) values ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
        sql = sql % (18, '属', pid, latin_name, shu_data[5], shu_data[6], latin_name, get_pinyin(shu_data[6]),
                     get_pinyin_prefix(shu_data[6]), datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                     1, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 0, bookid)
        shu_id = insertMysql(sql)


if __name__ == '__main__':
    #search_isbn()
    #print(html)
    read07Excel('C:/Users/dell/Desktop/国图数据比对（2016~2019）.xlsx')

    #get_page_html()

    #get_ke_context('http://www.efloras.org/florataxon.aspx?flora_id=2&taxon_id=10072',0)
    #get_ke_context('http://www.efloras.org/florataxon.aspx?flora_id=2&taxon_id=20790')
    #get_text_docx()
    #read07_excel('C:/Users/dell/Desktop/1.xlsx')
    #read07_excel_by('C:/Users/dell/Desktop/中国高等植物图鉴 (2).xlsx')

    #deal_with_gdcstj(189,'Arivela',193837)

    # data_count = select('select count(*) from zhiwu2', 'zhiwu', True)
    # page_count = data_count[0]
    # page_size = 1000
    # page_index = page_count // page_size
    # if page_count % page_size != 0:
    #     page_index = page_index + 1
    # for index in range(0, page_index):
    #     sql = 'select * from zhiwu2 limit %s,%s'
    #     sql = sql % (index*page_size, page_size)
    #     print(sql)

    #deal_with_foc()

    #get_text_docx()




    #shu_data = get_name_existed('Diapensia')
    #mings = ['f','fsdf','fsdf1','fsdfs','fsdfs']
    #print(mings[2:len(mings)])
    # i = 0
    # datas = get_foc();
    # for data in datas:
    #     i = i + 1
    #     print(data)
    #     if i >= 10:break
    ##lists = get_foc_vol_list()
    ##print(lists)
    ##get_foc_volume_list(lists)
    #print(getHtml('http://flora.huh.harvard.edu/FloraData/002/Vol11/foc11-Preface.htm'))
    #print(get_page_html())
    #vol = 'http://www.efloras.org/browse.aspx?flora_id=2&start_taxon_id=103074,volume_page.aspx?volume_id=2002&flora_id=2'
    #print(re.findall(r"\d+\.?\d*",vol),get_max_number(vol))