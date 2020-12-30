import sys
import os
import time
from PyQt5.QtWidgets import QApplication, QProgressBar, QWidget, QPushButton, QFileDialog, QLabel, QFrame, QMessageBox, QInputDialog, QLineEdit
from PyQt5.QtCore import Qt
# from PyQt5.QtGui import QIcon
#from PyQt5.QtWidgets import
import shutil, fitz
import openpyxl
from PIL import Image


class MainUI(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    # 窗体GUI部分
    def init_ui(self):
        # 设置窗体大小
        self.setGeometry(500, 200, 350, 200)
        # 设置固定大小
        self.setFixedSize(500, 370)
        self.setWindowTitle('pdf图片提取程序')
        #self.setWindowIcon(QIcon('icon/769160.png'))

        # 设置软件过期时间
        data = '2020-8-21 13:50:00'
        data_array = time.strptime(data, "%Y-%m-%d %H:%M:%S")
        timeStamp = int(time.mktime(data_array))
        # print(timeStamp)
        # 判断软件是否过期
        if time.time() > timeStamp:
            QMessageBox.warning(self, '', '软件已过期，请联系作者', QMessageBox.Yes)
        else:
            # 进度条
            self.pbar = QProgressBar(self)
            # 进图条位置及大小
            self.pbar.setGeometry(20, 200, 480, 10)

            # todo 按钮
            # 选择文件按钮
            self.btn_select_file = QPushButton('选择pdf文件', self)
            self.btn_select_file.move(20, 226)
            self.btn_select_file.setFixedSize(125, 30)
            self.btn_select_file.clicked.connect(self.file_dialog)
            # 输出文件按钮
            self.btn_output_path = QPushButton('选择输出文件夹', self)
            self.btn_output_path.move(20, 265)
            self.btn_output_path.setFixedSize(125, 30)
            self.btn_output_path.clicked.connect(self.output_dialog)
            # 开始按钮
            self.btn = QPushButton('开始', self)
            # 创建按钮并移动
            self.btn.move(150, 310)
            self.btn.setFixedSize(200, 40)
            # 点击按钮，连接事件函数
            self.btn.clicked.connect(self.btn_action)

            # todo 标签
            # 文件路径标签
            self.lab_select_path = QLabel('文件路径', self)
            self.lab_select_path.move(150, 225)
            self.lab_select_path.setFixedSize(320, 30)
            self.lab_select_path.setFrameShape(QFrame.Box)
            self.lab_select_path.setFrameShadow(QFrame.Raised)
            # 输出标签
            self.lab_output_path = QLabel('文件路径', self)
            self.lab_output_path.move(150, 265)
            self.lab_output_path.setFixedSize(320, 20)
            self.lab_output_path.setFrameShape(QFrame.Box)
            self.lab_output_path.setFrameShadow(QFrame.Raised)
            # 说明标签
            content = '说明:\n    选择需要提取图片的pdf；选择输出结果文件夹。\n'
            self.description_lab = QLabel(content, self)
            self.description_lab.move(20, 10)
            self.description_lab.setFixedSize(450, 100)
            self.description_lab.setAlignment(Qt.AlignTop)
            self.description_lab.setFrameShape(QFrame.Box)
            self.description_lab.setFrameShadow(QFrame.Raised)
            # 自动换行
            # self.description_lab.adjustSize()
            self.description_lab.setWordWrap(True)

            self.step = 0

            # 显示
            self.show()

    # 按钮点击
    def btn_action(self):
        if self.btn.text() == '完成':
            self.close()
        else:
            file_path = '{}'.format(self.lab_select_path.text())
            output_path = '{}'.format(self.lab_output_path.text())
            password = ('123', True)

            #ok = QInputDialog.getText(self, "Noza", "请输入密码：", QLineEdit.Password, '')
            #if ok == password:
            if self.btn.text() == '开始':
                if not os.path.exists(file_path):
                    QMessageBox.warning(self, '', '请选择路径', QMessageBox.Yes)
                elif not os.path.exists(output_path):
                    QMessageBox.warning(self, '', '请选择输出路径', QMessageBox.Yes)
                else:
                    self.btn.setText('程序进行中')
                    self.run()
            #else:
            #    QMessageBox.warning(self, '', '密码不正确！！！', QMessageBox.Yes)

    # 选择输入文件路径
    def file_dialog(self):
        # './'表示当前路径
        #path = QFileDialog.getExistingDirectory(self, '选取文件', './', 'pdf文件(*.pdf)')
        path = QFileDialog.getOpenFileName(self, '选取文件', './', 'pdf文件(*.pdf)')
        print(path)
        # 标签框显示文本路径
        self.lab_select_path.setText(path[0])
        # 自动调整标签框大小
        self.lab_select_path.adjustSize()

    # 选择输出路径
    def output_dialog(self):
        path = QFileDialog.getExistingDirectory(self, '选取文件夹', './')

        # 标签框显示文本路径
        self.lab_output_path.setText(path)
        # 自动调整标签框大小
        self.lab_output_path.adjustSize()

    def set_progerss_bar(self, num):
        '''
        设置进图条函数
        :param num: 进度条进度（整数）
        :return:
        '''
        self.step = num
        self.pbar.setValue(self.step)

    # 业务逻辑
    def run(self):
        # todo  文件路径
        file_path = '{}'.format(self.lab_select_path.text())
        output_path = '{}'.format(self.lab_output_path.text())

        # todo
        def get_file_name(file_dir):
            '''
            获取指定目录下所有文件名称
            :param file_dir:指定目录
            :return:返回文件名列表
            '''
            for root, dirs, files in os.walk(file_dir):
                # return root#当前目录路径
                # return dirs#当前路径下所有子目录
                return files  # 当前路径下所有非目录子文件

        #self.set_progerss_bar(10)

        def output2excel(file_dir, output_path):
            '''
            把文件夹下的文件名称输出到文件目录
            :param file_dir: 文件目录
            :return:
            '''

            # 获取文件目录下所有文件名，存入data列表
            data = get_file_name(file_dir)

            self.set_progerss_bar(50)

            # 把data输出到该目录下，并以目录名保存为excel格式
            wb = openpyxl.Workbook()
            sheet = wb.active
            # 设置表名为文件目录名
            sheet.title = '生成结果'
            for i in range(1, len(data) + 1):
                sheet['A{}'.format(i)] = data[i - 1]

            self.set_progerss_bar(80)

            wb.save('{0}/生成结果.xlsx'.format(output_path))

        def lookup_matrix(page, imgname):
            """Return the transformation matrix for an image name.

            Args:
                :page: the PyMuPDF page object
                :imgname: the image reference name, must equal the name in the
                    list doc.getPageImageList(page.number).
            Returns:
                concatenated matrices preceeding the image invocation.

            Notes:
                We are looking up "/imgname Do" in the concatenated /Contents of the
                page first. If not found, also look it up in the streams of any
                Form XObjects of the page. If still not found, return the zero matrix.
            """
            doc = page.parent  # get the PDF document
            if not doc.isPDF:
                raise ValueError("not PDF")

            page._cleanContents()  # sanitize image invocation matrices
            xref = page._getContents()[0]  # the (only) contents object
            cont = doc._getXrefStream(xref)  # the contents object
            cont = cont.replace(b"/", b" /")  # prepend slashes with a space
            # split this, ignoring white spaces
            cont = cont.split()

            imgnm = bytes("/" + imgname, "utf8")
            if imgnm in cont:
                idx = cont.index(imgnm)  # the image name is found here
            else:  # not in page /contents, so look in Form XObjects
                cont = None
                xreflist = doc._getPageInfo(page.number, 3)  # XObject xrefs
                for item in xreflist:
                    cont = doc._getXrefStream(item[0]).split()
                    if imgnm not in cont:
                        cont = None
                        continue
                    idx = cont.index(imgnm)  # image name found here
                    break

            if cont is None:  # safeguard against inconsistencies
                return fitz.Matrix()

            # list of matrices preceeding image invocation command.
            # not really required, because clean contents has concatenated those
            mat_list = []
            while idx >= 0:  # start value is "/Image Do" location
                if cont[idx] == b"q":  # finished at leading stacking command
                    break
                if cont[idx] == b"cm":  # encountered a matrix command
                    mat = cont[idx - 6: idx]  # list of the 6 matrix values
                    l = list(map(float, mat))  # make them floats
                    mat_list.append(fitz.Matrix(l))  # append fitz matrix
                    idx -= 6  # step backwards 6 entries
                else:
                    idx -= 1  # step backwards

            l = len(mat_list)
            if l == 0:  # safeguard against unusual situations
                return fitz.Matrix()  # the zero matrix

            mat = fitz.Matrix(1,
                              1)  # concatenate encountered matrices to this one
            for m in reversed(mat_list):
                mat *= m

            return mat

        def get_image_rotation(pic_path, doc, page, image_list, cols):
            short = lambda x: round(x, 4)
            i = 1
            #print(page.getText())
            for item in page.getImageList():  # loop through all images on the page
                #print(item)
                row = []
                row.append(page.number)
                img = item[7]  # we need the image reference name
                matrix = lookup_matrix(page, img)  # find matrix for the image
                if not bool(matrix):  # no display command  found for image
                    print("Image '%s' not found on page %i." % (img, page.number))
                    continue
                matrix = fitz.Matrix(tuple(map(short, matrix)))
                print(matrix)
                # print(page.getImageBbox(item))
                image_name = os.path.join(pic_path,
                                          "P{}_{}.png".format(page.number + 1,
                                                              i))
                pix = fitz.Pixmap(doc, item[0])
                rot1 = 0
                if min(matrix.a, matrix.d) > 0 and matrix.b == matrix.c == 0:
                    rot = "0 deg"
                elif matrix.a == matrix.d == 0:
                    if matrix.b > 0 and matrix.c < 0:
                        rot = "90 deg"
                    elif matrix.b < 0 and matrix.c > 0:
                        rot = "-90 deg"
                    else:
                        rot = "unknown"
                elif min(matrix.a, matrix.d) < 0 and matrix.b == matrix.c == 0:
                    rot = "180 deg"
                    rot1 = 180
                else:
                    rot = "unknown"
                print("Page %i / %i, image '%s', rotation: %s." % (page.number, doc.pageCount, img, rot))
                if pix.n - pix.alpha < 4:
                    pix.writePNG(image_name)
                else:  # 否则先转换CMYK
                    pix0 = fitz.Pixmap(fitz.csRGB, pix)
                    pix0.writePNG(image_name)
                    pix0 = None
                if rot1 == 180:
                    image = Image.open(image_name)
                    out3 = image.transpose(Image.FLIP_TOP_BOTTOM)
                    out3.save(image_name)

                i = i + 1
                row.append(image_name)
                cols.append(row)

        def create_xls(datas, file_path):
            """
            :param file_path: 新建并保存的xls路径及名称
            :param sheet_name: sheet名
            :param head_name: 表格首行
            :param datas: 表格数据
            :return:
            """
            # file_path = '/Users/mac5318/Downloads/demo.xls'
            wb = openpyxl.Workbook()
            ws = wb.active
            # ws1 = wb.get_sheet_by_name('sheet1')
            ws.title = 'sheet1'
            # ws2 = wb.create_sheet('sheet1')

            for data in datas:
                ws.append(data)
            wb.save(file_path)

        def extract_pdf(pdf_path, pic_path):
            doc = fitz.open(pdf_path)
            cols = []
            row = []
            row.append('页码')
            row.append('路径')
            cols.append(row)
            for page in doc:
                #dict_text = page.getText('dict')  ## {'type': 0, 'bbox': (0,0,0,0), 'lines': [{'wmode': 0, 'dir': (1.0, 0.0), 'bbox':...
                list_image = page.getImageList()
                #print(list_image)
                #print(dict_text['blocks'])
                # if page.number > 80:
                #     break
                get_image_rotation(pic_path, doc, page, list_image, cols)
                self.set_progerss_bar(page.number/doc.pageCount*100)
            create_xls(cols, os.path.join(pic_path, "page.xlsx"))

        shutil.copyfile(file_path, os.path.join(output_path, os.path.splitext(os.path.basename(file_path))[0] + '.pdf'))
        #output2excel(file_path, output_path)
        extract_pdf(file_path, output_path)
        self.set_progerss_bar(100)
        self.btn.setText('完成')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    pbar = MainUI()
    sys.exit(app.exec_())
    # file_path = '/Users/mac5318/Downloads/27juan.pdf'
    # output_path = '/Users/mac5318/Downloads/27juan'
    # print(os.path.splitext(os.path.basename(file_path))[0] + '.pdf')
